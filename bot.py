import asyncio
import os
import pytz
from datetime import datetime, timedelta

from aiogram import Bot, Dispatcher, F
from aiogram.filters import CommandStart
from aiogram.types import Message, CallbackQuery
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl import Workbook, load_workbook

# ================== НАСТРОЙКИ ==================
TOKEN = "8678946305:AAGSHg3mGWiIy8K2jEChMmcghppKPNhsGXE"
ADMIN_ID = 980044682
EXCEL_FILE = "answers.xlsx"

dp = Dispatcher()
users = {}

# ================== ТЕКСТЫ ==================
SCHOOL_BLOCK_TEXT = (
    "🎓 **Visions.kz** — школа практического обучения по Scrum и фасилитации.\n\n"
    "Мы учим выстраивать эффективную работу команд и проводить результативные встречи.\n\n"
    "📌 **Основные направления:**\n"
    "• Курс «Scrum»\n"
    "• Курс «Фасилитация»\n\n"
    "💼 **Что делает менеджер:**\n"
    "• Ищет потенциальных клиентов\n"
    "• Консультирует по программам обучения\n"
    "• Объясняет ценность курса и пользу обучения\n"
    "• Сопровождает клиента до оплаты\n\n"
    "🖥 **Формат работы:**\n"
    "• Удалённо\n\n"
    "💳**Условия оплаты:**\n"
    "Мы предлагаем сотрудничество, где оплата производится только за выполненные задачи."
    "Нет фиксированного оклада, но есть возможность заработать больше, "
    "если вы готовы работать на результат.\n\n"
    "Нажмите кнопку ниже, чтобы продолжить ⬇️"
)

FORM_STEPS = [
    ("fio", "1/5) Укажите ваше ФИО."),
    ("city", "2/5) Ваш город?"),
    ("sales_exp", "3/5) Сколько лет опыта в продажах?"),
    ("why_us", "4/5) Почему хотите работать именно в нашей школе?"),
    ("hours", "5/5) Сколько часов готовы уделять работе?")
]

HEADERS = [
    "Дата",
    "Telegram ID",
    "Username",
    "ФИО",
    "Город",
    "Опыт в продажах",
    "Почему Visions",
    "Часы в неделю",
    "Scrum ответ",
    "Facilitation ответ"
]


# ================== ВСПОМОГАТЕЛЬНОЕ ==================
def get_next_interview_datetime(now: datetime) -> datetime:
    # Среда = 2
    days_ahead = 2 - now.weekday()
    if days_ahead <= 0:
        days_ahead += 7
    interview = now + timedelta(days=days_ahead)
    return interview.replace(hour=12, minute=30, second=0, microsecond=0)


# import gspread
# from google.oauth2.service_account import Credentials
#
# SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]
# creds = Credentials.from_service_account_file(
#     "service_account.json",
#     scopes=SCOPES
# )
#
# gc = gspread.authorize(creds)
# sheet = gc.open("Visions.kz Applicants").sheet1
#
#
# def save_to_google(user_id, username, a):
#     sheet.append_row([
#         datetime.now().strftime("%Y-%m-%d %H:%M"),
#         user_id,
#         username,
#         a["fio"],
#         a["city"],
#         a["sales_exp"],
#         a["why_us"],
#         a["hours"],
#         a.get("scrum_answer"),
#         a.get("fac_answer"),
#     ])

def save_to_excel(user_id, username, a):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([
        datetime.now().strftime("%Y-%m-%d %H:%M"),
        user_id,
        username,
        a["fio"],
        a["city"],
        a["sales_exp"],
        a["why_us"],
        a["hours"],
        a.get("scrum_answer"),
        a.get("fac_answer"),
    ])
    wb.save(EXCEL_FILE)


# ================== КЛАВИАТУРЫ ==================
def start_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="▶️ Начать отбор", callback_data="start")
    return kb.as_markup()


def continue_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Продолжить", callback_data="continue")
    return kb.as_markup()


def yes_no_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="Да", callback_data="yes")
    kb.button(text="Нет", callback_data="no")
    kb.adjust(2)
    return kb.as_markup()


def yes_no_kb1():
    kb = InlineKeyboardBuilder()
    kb.button(text="Да", callback_data="yes1")
    kb.button(text="Нет", callback_data="no1")
    kb.adjust(2)
    return kb.as_markup()


def scrum_kb():
    kb = InlineKeyboardBuilder()
    kb.button(text="О создании ценности для клиента", callback_data="scrum_value")
    kb.button(text="О жёстком контроле сроков", callback_data="scrum_deadline")
    kb.button(text="О подробной документации", callback_data="scrum_docs")
    kb.adjust(1)
    return kb.as_markup()


def fac_kb():
    kb = InlineKeyboardBuilder()
    kb.button(
        text="Помогать группе достичь результата",
        callback_data="fac_help"
    )
    kb.button(text="Принимать решения за команду", callback_data="fac_decide")
    kb.button(text="Контролировать и оценивать участников", callback_data="fac_control")
    kb.adjust(1)
    return kb.as_markup()


# ================== ХЭНДЛЕРЫ ==================
@dp.message(CommandStart())
async def start(message: Message):
    await message.answer(
        "Привет 👋\n"
        "Это отбор на позицию менеджера по продажам.\n"
        "Прохождение займёт 7–10 минут.",
        reply_markup=start_kb()
    )


@dp.callback_query(F.data == "start")
async def start_selection(cb: CallbackQuery):
    await cb.answer()
    uid = cb.from_user.id

    if uid in users:
        state = users[uid]["state"]

        if state == "started":
            await cb.message.answer(
                "⏳ Вы уже начали отбор.\n"
                "Пожалуйста, завершите его."
            )
            return

        if state == "interview":
            await cb.message.answer(
                "🎉 Вы уже прошли отбор!\n"
                "Наш менеджер свяжется с вами."
            )
            return

        if state == "closed":
            await cb.message.answer(
                "❌ Отбор для вас завершён.\n"
                "Если появится новая возможность — мы сообщим."
            )
            return

    # если пользователя ещё нет — создаём
    users[uid] = {
        "state": "started",
        "stage": "form",
        "step": 0,
        "answers": {},
        "started_at": datetime.now(),
        "last_activity": datetime.now(),
        "interview_at": None,
        "rem_24h": False,
        "closed_48h": False,
        "rem_1d_interview": False,
        "rem_1h_interview": False,
    }

    await cb.message.answer(SCHOOL_BLOCK_TEXT, reply_markup=continue_kb(), parse_mode="Markdown")


@dp.callback_query(F.data == "continue")
async def ask_fit(cb: CallbackQuery):
    await cb.answer()
    await cb.message.answer(
        "Подходит ли вам формат работы и сфера обучения?",
        reply_markup=yes_no_kb()
    )


@dp.callback_query(F.data == "no")
async def finish_no(cb: CallbackQuery):
    await cb.answer()

    await cb.message.edit_reply_markup(reply_markup=None)
    await cb.message.answer("Спасибо за проявленный интерес к Visions.kz.\n"
                            "Желаем вам успехов в дальнейшей профессиональной деятельности.")


@dp.callback_query(F.data == "yes")
async def start_form(cb: CallbackQuery):
    await cb.answer()

    await cb.message.edit_reply_markup(reply_markup=None)

    await cb.message.answer("Отлично! Начнём анкету.\n✏️Данные записываем.")
    await ask_question(cb.message, cb.from_user.id)


@dp.callback_query(F.data == "yes1")
async def confirm_interview(cb: CallbackQuery):
    await cb.answer()
    await cb.message.edit_reply_markup(reply_markup=None)

    await cb.message.answer(
        "✅ Отлично! Вот ссылка на Zoom:\n\n"
        "🔗 https://us02web.zoom.us/j/82306121295?pwd=PzHSa5H97mPJInOmJ2szx45A9GEPpL.1\n\n"
        "🕧 Среда, 12:30\n"
        "Пожалуйста, подключитесь вовремя."
    )


async def ask_question(message: Message, user_id: int):
    step = users[user_id]["step"]
    await message.answer(FORM_STEPS[step][1])


@dp.message(F.text)
async def handle_text(message: Message):
    uid = message.from_user.id
    if uid not in users:
        return

    u = users[uid]
    u["last_activity"] = datetime.now()

    if u["stage"] == "form":
        key = FORM_STEPS[u["step"]][0]
        u["answers"][key] = message.text
        u["step"] += 1

        if u["step"] >= len(FORM_STEPS):
            u["stage"] = "scrum"
            await message.answer(
                "📺 Посмотрите видео о Scrum\n"
                "https://youtu.be/taG2vXriZbM\n\n"
                "Это поможет лучше понять суть подхода.\n\n"
                "<b>После просмотра видео, выберите верный ответ:</b>\n"

                "О чём Scrum в первую очередь?\n",
                reply_markup=scrum_kb(),
                parse_mode="HTML",
                disable_web_page_preview=True
            )
        else:
            await ask_question(message, uid)


@dp.callback_query(F.data.startswith("scrum_"))
async def scrum_answer(cb: CallbackQuery):
    await cb.answer()
    u = users[cb.from_user.id]
    u["answers"]["scrum_answer"] = cb.data
    u["stage"] = "fac"

    await cb.message.answer(
        "📺 Посмотрите видео про Фасилитацию:\n"
        "https://youtube.com/shorts/D3xq0GF5HI0\n\n"
        "<b>После просмотра видео, выберите верный ответ:</b>\n"
        "Вопрос: В чём основная роль фасилитатора?",
        reply_markup=fac_kb(),
        parse_mode="HTML",
        disable_web_page_preview=True
    )


@dp.callback_query(F.data.startswith("fac_"))
async def fac_answer(cb: CallbackQuery):
    await cb.answer()
    u = users[cb.from_user.id]
    u["answers"]["fac_answer"] = cb.data

    u["state"] = "interview"
    u["interview_at"] = get_next_interview_datetime(datetime.now())

    save_to_excel(
        cb.from_user.id,
        cb.from_user.username or "",
        u["answers"]
    )

    await cb.message.answer(
        "🎉 Поздравляем! Вы прошли отбор.\n\n"
        "Подпишитесь на каналы школы, чтобы узнать больше o visions.kz:\n"
        "📸 Instagram: https://instagram.com/visions.kz\n"
        "📲 Telegram: https://t.me/visionskz\n\n"
        "Приглашаем вас на онлайн-встречу:\n"
        "🗓 Среда, 12:30\n\n"
        "Подтвердите участие 👇🏻",
        reply_markup=yes_no_kb1(),
        parse_mode="HTML",
        disable_web_page_preview=True
    )


# ================== НАПОМИНАНИЯ ==================
async def reminder_worker(bot: Bot):
    while True:
        tz = pytz.timezone("Asia/Almaty")
        now = datetime.now(tz)

        for uid, u in list(users.items()):
            # 1–2. Не закончил отбор
            if u["state"] == "started":
                delta = now - u["started_at"]

                if delta >= timedelta(hours=24) and not u["rem_24h"]:
                    await bot.send_message(
                        uid,
                        "⏰ Напоминание!\n"
                        "Вы начали отбор, но не завершили его.\n"
                        "Это займёт ещё 5–7 минут 😊"
                    )
                    u["rem_24h"] = True

                if delta >= timedelta(hours=48) and not u["closed_48h"]:
                    await bot.send_message(
                        uid,
                        "❌ Время на прохождение отбора истекло.\n"
                        "Если интерес останется — вы сможете пройти его позже."
                    )
                    u["state"] = "closed"
                    u["closed_48h"] = True

            # 3–4. Напоминания про интервью
            if u["state"] == "interview":
                interview_at = u["interview_at"]
                delta = interview_at - now

                # Напоминание за 1 день
                if timedelta(0) <= delta <= timedelta(days=1) and not u["rem_1d_interview"]:
                    await bot.send_message(
                        uid,
                        "📅 Напоминание!\n"
                        "Завтра у вас интервью с Visions.kz\n"
                        "🕧 Среда, 12:30"
                    )
                    u["rem_1d_interview"] = True

                # Напоминание за 1 час
                if timedelta(0) <= delta <= timedelta(hours=1) and not u["rem_1h_interview"]:
                    await bot.send_message(
                        uid,
                        "⏰ Интервью начнётся через 1 час!\n"
                        "🕧 Сегодня в 12:30\n"
                        "Пожалуйста, будьте на связи."
                    )
                    u["rem_1h_interview"] = True

        await asyncio.sleep(300)


# ================== ЗАПУСК ==================
async def main():
    bot = Bot(TOKEN)
    asyncio.create_task(reminder_worker(bot))
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())